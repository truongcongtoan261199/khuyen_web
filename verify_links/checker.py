import re
import sys
import time
import argparse
import socket
from urllib.parse import urlparse, urljoin
import urllib.request
import urllib.error
import urllib.parse
from html.parser import HTMLParser

# Cố gắng import socks để hỗ trợ proxy SOCKS
try:
    import socks
    from sockshandler import SocksiPyHandler

    HAS_SOCKS = True
except ImportError:
    HAS_SOCKS = False

# ── ANSI colours ──────────────────────────────────────────────────────────────
GREEN = "\033[92m"
YELLOW = "\033[93m"
RED = "\033[91m"
CYAN = "\033[96m"
BOLD = "\033[1m"
DIM = "\033[2m"
RESET = "\033[0m"


def ok(msg):
    return f"{GREEN}✓  {msg}{RESET}"


def warn(msg):
    return f"{YELLOW}⚠  {msg}{RESET}"


def blocked(msg):
    return f"{RED}✗  {msg}{RESET}"


def info(msg):
    return f"{CYAN}ℹ  {msg}{RESET}"


def dim(msg):
    return f"{DIM}{msg}{RESET}"


# ── HTTP helpers ──────────────────────────────────────────────────────────────

DEFAULT_UA = "Mozilla/5.0 (compatible; RobotsExclusionChecker/1.0)"


def get_opener(proxy_url: str = None):
    """Tạo urllib opener hỗ trợ Proxy (HTTP/SOCKS)."""
    handlers = []
    if proxy_url:
        parsed_proxy = urlparse(proxy_url)
        scheme = parsed_proxy.scheme.lower()

        if "socks" in scheme:
            if not HAS_SOCKS:
                print(
                    warn(
                        "Thư viện 'PySocks' chưa được cài đặt. Proxy SOCKS sẽ không hoạt động."
                    )
                )
            else:
                proxy_type = socks.SOCKS5 if "5" in scheme else socks.SOCKS4
                handlers.append(
                    SocksiPyHandler(
                        proxy_type, parsed_proxy.hostname, parsed_proxy.port
                    )
                )
        else:
            # Mặc định cho HTTP/HTTPS Proxy
            handlers.append(
                urllib.request.ProxyHandler({"http": proxy_url, "https": proxy_url})
            )

    return urllib.request.build_opener(*handlers)


def http_get(
    url: str, ua: str = DEFAULT_UA, proxy: str = None
) -> tuple[int, dict, str]:
    """Return (status_code, headers_dict, body_text)."""
    opener = get_opener(proxy)
    req = urllib.request.Request(url, headers={"User-Agent": ua})
    try:
        with opener.open(req, timeout=15) as resp:
            headers = dict(resp.getheaders())
            body = resp.read().decode("utf-8", errors="replace")
            return resp.status, headers, body
    except urllib.error.HTTPError as e:
        headers = dict(e.headers)
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            body = ""
        return e.code, headers, body
    except Exception as e:
        raise ConnectionError(f"Could not reach {url}: {e}") from e


# ══════════════════════════════════════════════════════════════════════════════
# 1. ROBOTS.TXT CHECKER
# ══════════════════════════════════════════════════════════════════════════════


def fetch_robots_txt(base_url: str, proxy: str = None) -> str:
    parsed = urlparse(base_url)
    robots_url = f"{parsed.scheme}://{parsed.netloc}/robots.txt"
    try:
        _, _, body = http_get(robots_url, proxy=proxy)
        return body
    except Exception:
        return ""


def parse_robots_groups(content: str) -> list[dict]:
    """Parse robots.txt into groups of {agents, rules, crawl_delay}."""
    groups = []
    current_agents: list[str] = []
    current_rules: list[dict] = []
    current_delay = None

    def flush():
        if current_agents:
            groups.append(
                {
                    "agents": list(current_agents),
                    "rules": list(current_rules),
                    "crawl_delay": current_delay,
                }
            )

    for line in content.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        colon = line.find(":")
        if colon == -1:
            continue
        key = line[:colon].strip().lower()
        val = line[colon + 1 :].strip()

        if key == "user-agent":
            if current_rules:
                flush()
                current_agents = [val]
                current_rules = []
                current_delay = None
            else:
                current_agents.append(val)
        elif key in ("allow", "disallow"):
            current_rules.append({"type": key, "path": val})
        elif key == "crawl-delay":
            try:
                current_delay = float(val)
            except ValueError:
                pass

    flush()
    return groups


def _make_pattern(path: str) -> re.Pattern:
    escaped = re.escape(path).replace(r"\*", ".*")
    if escaped.endswith(r"\$"):
        escaped = escaped[:-2] + "$"
    return re.compile("^" + escaped)


def check_robots_txt(
    url: str, groups: list[dict], user_agent: str = "Googlebot"
) -> dict:
    parsed = urlparse(url)
    path = parsed.path or "/"
    if parsed.query:
        path += "?" + parsed.query

    specific = [
        g for g in groups if user_agent.lower() in [a.lower() for a in g["agents"]]
    ]
    wildcard = [g for g in groups if "*" in g["agents"]]
    applicable = specific if specific else wildcard

    best_len = -1
    best_result = True
    best_rule = None
    best_agent = None
    best_delay = None

    for group in applicable:
        agent_label = ", ".join(group["agents"])
        if group["crawl_delay"] is not None:
            best_delay = group["crawl_delay"]
        for rule in group["rules"]:
            rpath = rule["path"]
            if rpath == "":
                if rule["type"] == "disallow" and best_len == -1:
                    best_len = 0
                    best_result = True
                    best_rule = rule
                    best_agent = agent_label
                continue
            try:
                if _make_pattern(rpath).match(path):
                    if len(rpath) > best_len:
                        best_len = len(rpath)
                        best_result = rule["type"] == "allow"
                        best_rule = rule
                        best_agent = agent_label
            except re.error:
                pass

    return {
        "allowed": best_result,
        "matched_rule": best_rule,
        "matched_agent": best_agent,
        "crawl_delay": best_delay,
        "status": "ok" if best_result else "blocked",
    }


# ══════════════════════════════════════════════════════════════════════════════
# 2. CANONICAL TAG CHECKER
# ══════════════════════════════════════════════════════════════════════════════


class _MetaParser(HTMLParser):
    def __init__(self, page_url: str):
        super().__init__()
        self.page_url = page_url
        self.canonical = None
        self.meta_robots: list[str] = []
        self._in_head = True

    def handle_starttag(self, tag, attrs):
        if not self._in_head:
            return
        attrs_d = dict(attrs)

        if tag == "link":
            rel = (attrs_d.get("rel") or "").lower()
            href = attrs_d.get("href") or ""
            if rel == "canonical" and href:
                self.canonical = urljoin(self.page_url, href)

        elif tag == "meta":
            name = (attrs_d.get("name") or "").lower()
            content = (attrs_d.get("content") or "").lower()
            if name in ("robots", "googlebot", "bingbot") and content:
                self.meta_robots.append(f'{attrs_d.get("name", "robots")}: {content}')

    def handle_endtag(self, tag):
        if tag == "head":
            self._in_head = False


def check_canonical(url: str, body: str) -> dict:
    parser = _MetaParser(url)
    parser.feed(body)
    canonical = parser.canonical
    if canonical is None:
        status = "warn"
        message = "No canonical tag found"
    elif canonical.rstrip("/") == url.rstrip("/"):
        status = "ok"
        message = "Self-referencing canonical (correct)"
    else:
        status = "warn"
        message = "Canonicalizes to a different URL"
    return {
        "canonical": canonical,
        "page_url": url,
        "status": status,
        "message": message,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 3. META ROBOTS CHECKER
# ══════════════════════════════════════════════════════════════════════════════


def check_meta_robots(body: str) -> dict:
    parser = _MetaParser("http://x")
    parser.feed(body)
    directives = parser.meta_robots

    issues = []
    for d in directives:
        val = d.lower()
        if "noindex" in val:
            issues.append("noindex")
        if "nofollow" in val:
            issues.append("nofollow")
        if "noarchive" in val:
            issues.append("noarchive")
        if "noimageindex" in val:
            issues.append("noimageindex")
        if "nosnippet" in val:
            issues.append("nosnippet")
        if "none" in val:
            issues.append("none (noindex + nofollow)")

    if not directives:
        status = "ok"
        message = "No meta robots tags found (default: index, follow)"
    elif "noindex" in issues or "none" in " ".join(issues):
        status = "blocked"
        message = "Page has noindex directive — will not appear in search results"
    elif issues:
        status = "warn"
        message = f"Restrictive directives detected: {', '.join(set(issues))}"
    else:
        status = "ok"
        message = "Meta robots present but no blocking directives"

    return {
        "directives": directives,
        "issues": list(set(issues)),
        "status": status,
        "message": message,
    }


# ══════════════════════════════════════════════════════════════════════════════
# 4. HTTP HEADER CHECKER  (X-Robots-Tag)
# ══════════════════════════════════════════════════════════════════════════════


def check_http_headers(headers: dict) -> dict:
    x_robots = [v for k, v in headers.items() if k.lower() == "x-robots-tag"]

    issues = []
    for val in x_robots:
        v = val.lower()
        for directive in (
            "noindex",
            "nofollow",
            "noarchive",
            "noimageindex",
            "nosnippet",
            "none",
        ):
            if directive in v:
                issues.append(directive)

    extra = {}
    for key in (
        "Content-Type",
        "Last-Modified",
        "Cache-Control",
        "ETag",
        "X-Frame-Options",
    ):
        for hk, hv in headers.items():
            if hk.lower() == key.lower():
                extra[key] = hv

    if not x_robots:
        status = "ok"
        message = "No X-Robots-Tag header found (no server-level blocking)"
    elif "noindex" in issues or "none" in issues:
        status = "blocked"
        message = "X-Robots-Tag contains noindex — page blocked at server level"
    elif issues:
        status = "warn"
        message = f"X-Robots-Tag has restrictive directives: {', '.join(set(issues))}"
    else:
        status = "ok"
        message = "X-Robots-Tag present but no blocking directives"

    return {
        "x_robots_tag": x_robots,
        "issues": list(set(issues)),
        "extra_headers": extra,
        "status": status,
        "message": message,
    }


# ══════════════════════════════════════════════════════════════════════════════
# OVERALL STATUS
# ══════════════════════════════════════════════════════════════════════════════


def overall_status(results: dict) -> str:
    statuses = [v.get("status") for v in results.values() if isinstance(v, dict)]
    if "blocked" in statuses:
        return "blocked"
    if "warn" in statuses:
        return "warn"
    return "ok"


# ══════════════════════════════════════════════════════════════════════════════
# DISPLAY (single URL mode)
# ══════════════════════════════════════════════════════════════════════════════


def traffic_light(status: str) -> str:
    return {
        "ok": f"{GREEN}● ALL CLEAR{RESET}",
        "warn": f"{YELLOW}● POTENTIAL ISSUE{RESET}",
        "blocked": f"{RED}● BLOCKED{RESET}",
    }.get(status, status)


def section(title: str):
    print(f"\n{BOLD}{CYAN}{'─'*60}{RESET}")
    print(f"{BOLD}  {title}{RESET}")
    print(f"{CYAN}{'─'*60}{RESET}")


def print_results(url: str, user_agent: str, results: dict):
    overall = overall_status(results)

    print(f"\n{'═'*60}")
    print(f"{BOLD}  Robots Exclusion Checker{RESET}")
    print(f"{'═'*60}")
    print(f"  URL:        {url}")
    print(f"  User-agent: {user_agent}")
    print(f"  Status:     {traffic_light(overall)}")

    section("1 · Robots.txt")
    r = results["robots_txt"]
    if r["status"] == "ok":
        print(ok("Crawling allowed"))
    else:
        print(blocked("Crawling BLOCKED by robots.txt"))
    rule = r["matched_rule"]
    if rule:
        print(dim(f"     Rule matched : {rule['type']}: {rule['path'] or '(all)'}"))
        print(dim(f"     Agent block  : {r['matched_agent']}"))
    else:
        print(dim("     No matching rule — allowed by default"))
    if r["crawl_delay"]:
        print(warn(f"Crawl-delay directive found: {r['crawl_delay']}s"))

    section("2 · Canonical Tag")
    c = results["canonical"]
    fn = {"ok": ok, "warn": warn, "blocked": blocked}[c["status"]]
    print(fn(c["message"]))
    if c["canonical"]:
        print(dim(f"     <link rel=\"canonical\" href=\"{c['canonical']}\">"))
    else:
        print(dim("     No canonical tag in <head>"))

    section("3 · Meta Robots")
    m = results["meta_robots"]
    fn = {"ok": ok, "warn": warn, "blocked": blocked}[m["status"]]
    print(fn(m["message"]))
    if m["directives"]:
        for d in m["directives"]:
            print(dim(f"     {d}"))
    if m["issues"]:
        for issue in m["issues"]:
            print(warn(f"Directive detected: {issue}"))

    section("4 · HTTP Headers (X-Robots-Tag)")
    h = results["http_headers"]
    fn = {"ok": ok, "warn": warn, "blocked": blocked}[h["status"]]
    print(fn(h["message"]))
    if h["x_robots_tag"]:
        for val in h["x_robots_tag"]:
            print(dim(f"     X-Robots-Tag: {val}"))
    if h["issues"]:
        for issue in h["issues"]:
            print(warn(f"Directive detected: {issue}"))
    if h["extra_headers"]:
        print(dim("     Other headers:"))
        for k, v in h["extra_headers"].items():
            print(dim(f"       {k}: {v}"))

    print(f"\n{'═'*60}\n")


# ══════════════════════════════════════════════════════════════════════════════
# SINGLE URL RUN
# ══════════════════════════════════════════════════════════════════════════════


def run(url: str, user_agent: str = "Googlebot", proxy: str = None):
    if not url.startswith("http"):
        url = "https://" + url

    print(info(f"Fetching page: {url}"))
    if proxy:
        print(dim(f"  Using proxy: {proxy}"))

    try:
        status_code, headers, body = http_get(url, user_agent, proxy)
        print(dim(f"  HTTP {status_code}"))

        print(info("Fetching robots.txt…"))
        robots_content = fetch_robots_txt(url, proxy)
        groups = parse_robots_groups(robots_content) if robots_content else []

        results = {
            "robots_txt": check_robots_txt(url, groups, user_agent),
            "canonical": check_canonical(url, body),
            "meta_robots": check_meta_robots(body),
            "http_headers": check_http_headers(headers),
        }

        print_results(url, user_agent, results)
        return results
    except Exception as e:
        print(blocked(f"Error: {e}"))
        return None


# ══════════════════════════════════════════════════════════════════════════════
# BULK MODE
# ══════════════════════════════════════════════════════════════════════════════


def _bulk_mode(
    input_path: str, ua: str, workers: int, out_path: str, proxy: str = None
):
    """Check many URLs from Excel and write results to Excel."""
    from datetime import datetime
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from threading import Lock
    import pandas as pd
    from openpyxl.styles import Font, PatternFill

    # ── robots.txt cache per domain ───────────────────────────────────────────
    _robots_cache: dict[str, list] = {}
    _cache_lock = Lock()

    def get_robots_groups(url: str) -> list:
        parsed = urllib.parse.urlparse(url)
        domain = f"{parsed.scheme}://{parsed.netloc}"
        with _cache_lock:
            if domain in _robots_cache:
                return _robots_cache[domain]
        try:
            body = fetch_robots_txt(url, proxy)
            groups = parse_robots_groups(body) if body else []
        except Exception:
            groups = []
        with _cache_lock:
            _robots_cache[domain] = groups
        return groups

    def check_url_bulk(url: str) -> dict:
        url = url.strip()
        if not url.startswith("http"):
            url = "https://" + url

        try:
            status, headers, body = http_get(url, ua, proxy)
        except Exception as e:
            return {
                "url": url,
                "overall": "ERROR",
                "http_status": "-",
                "robots_txt": "ERROR",
                "canonical": "ERROR",
                "meta_robots": "ERROR",
                "x_robots_tag": "ERROR",
                "notes": str(e),
            }

        groups = get_robots_groups(url)
        r_txt = check_robots_txt(url, groups, ua)
        r_can = check_canonical(url, body)
        r_meta = check_meta_robots(body)
        r_xhdr = check_http_headers(headers)

        # Overall: BLOCKED if any check is blocked, WARN if any is warn, else PASS
        statuses = [
            r_txt["status"],
            r_can["status"],
            r_meta["status"],
            r_xhdr["status"],
        ]
        if "blocked" in statuses:
            overall = "BLOCKED"
        elif "warn" in statuses:
            overall = "WARN"
        else:
            overall = "PASS"

        # Build notes for non-ok checks
        notes_parts = []
        if r_txt["status"] != "ok":
            rule = r_txt["matched_rule"]
            notes_parts.append(
                f"robots.txt: {rule['type']} {rule['path']}"
                if rule
                else "robots.txt: blocked"
            )
        if r_can["status"] != "ok":
            notes_parts.append(f"canonical: {r_can['message']}")
        if r_meta["status"] != "ok":
            notes_parts.append(f"meta_robots: {r_meta['message']}")
        if r_xhdr["status"] != "ok":
            notes_parts.append(f"x_robots_tag: {r_xhdr['message']}")

        return {
            "url": url,
            "overall": overall,
            "http_status": status,
            "robots_txt": r_txt["status"].upper(),
            "canonical": r_can["status"].upper(),
            "meta_robots": r_meta["status"].upper(),
            "x_robots_tag": r_xhdr["status"].upper(),
            "canonical_url": r_can["canonical"] or "",
            "notes": " | ".join(notes_parts) if notes_parts else "all clear",
        }

    # ── Load URLs ─────────────────────────────────────────────────────────────
    try:
        df_in = pd.read_excel(input_path, dtype=str)
    except Exception as e:
        sys.exit(f"Error reading Excel file: {e}")

    if "link" not in df_in.columns:
        sys.exit(f"Column 'link' not found. Available: {list(df_in.columns)}")

    urls = [u.strip() for u in df_in["link"].dropna().tolist() if str(u).strip()]
    if not urls:
        sys.exit("No URLs found in 'link' column")

    print(f"✓ Loaded {len(urls)} URLs from '{input_path}'")

    # ── Concurrent processing ─────────────────────────────────────────────────
    total = len(urls)
    results = [None] * total
    done = 0
    start_time = time.time()

    print(f"\n{'='*70}")
    proxy_msg = f"  |  proxy={proxy}" if proxy else ""
    print(f"Checking {total} URLs  |  workers={workers}  |  user-agent={ua}{proxy_msg}")
    print(f"{'='*70}\n")

    with ThreadPoolExecutor(max_workers=workers) as pool:
        future_to_idx = {
            pool.submit(check_url_bulk, url): i for i, url in enumerate(urls)
        }
        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            try:
                results[idx] = future.result()
            except Exception as e:
                results[idx] = {
                    "url": urls[idx],
                    "overall": "ERROR",
                    "http_status": "-",
                    "robots_txt": "ERROR",
                    "canonical": "ERROR",
                    "meta_robots": "ERROR",
                    "x_robots_tag": "ERROR",
                    "canonical_url": "",
                    "notes": str(e),
                }
            done += 1
            elapsed = time.time() - start_time
            rate = done / elapsed if elapsed > 0 else 0
            eta = (total - done) / rate if rate > 0 else 0
            pct = done / total * 100
            bar = ("█" * int(pct / 2)).ljust(50)
            print(
                f"\r[{bar}] {done}/{total} ({pct:.0f}%)  {rate:.1f}/s  ETA {eta:.0f}s",
                end="",
                flush=True,
            )

    print()

    # ── Write Excel output ────────────────────────────────────────────────────
    columns = [
        "url",
        "overall",
        "http_status",
        "robots_txt",
        "canonical",
        "meta_robots",
        "x_robots_tag",
        "canonical_url",
        "notes",
    ]
    df_out = pd.DataFrame(results, columns=columns)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Results")
        ws = writer.sheets["Results"]

        # Auto-width columns
        for col in ws.columns:
            try:
                max_len = max(len(str(c.value or "")) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 80)
            except Exception:
                pass

        # Header bold + background
        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Color overall column (B)
        fill_pass = PatternFill("solid", fgColor="C6EFCE")
        fill_warn = PatternFill("solid", fgColor="FFEB9C")
        fill_blocked = PatternFill("solid", fgColor="FFC7CE")
        fill_error = PatternFill("solid", fgColor="D9D9D9")

        font_pass = Font(color="276221", bold=True, name="Arial")
        font_warn = Font(color="9C6500", bold=True, name="Arial")
        font_blocked = Font(color="9C0006", bold=True, name="Arial")
        font_error = Font(color="595959", bold=True, name="Arial")

        status_style = {
            "PASS": (fill_pass, font_pass),
            "WARN": (fill_warn, font_warn),
            "BLOCKED": (fill_blocked, font_blocked),
            "ERROR": (fill_error, font_error),
        }

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            overall_cell = row[1]  # column B = "overall"
            style = status_style.get(str(overall_cell.value), None)
            if style:
                overall_cell.fill, overall_cell.font = style
            # Apply Arial to all cells
            for cell in row:
                cell.font = Font(
                    name="Arial",
                    bold=cell.font.bold if cell.font else False,
                    color=(
                        cell.font.color.rgb
                        if cell.font
                        and cell.font.color
                        and cell.font.color.type == "rgb"
                        else "000000"
                    ),
                )

        # ── Summary sheet ─────────────────────────────────────────────────────
        elapsed = time.time() - start_time
        ws2 = writer.book.create_sheet("Summary", 0)

        summary_data = [
            ("Metric", "Value"),
            ("Total URLs", total),
            ("PASS", sum(1 for r in results if r and r["overall"] == "PASS")),
            ("WARN", sum(1 for r in results if r and r["overall"] == "WARN")),
            ("BLOCKED", sum(1 for r in results if r and r["overall"] == "BLOCKED")),
            ("ERROR", sum(1 for r in results if r and r["overall"] == "ERROR")),
            (
                "Pass Rate",
                f"{sum(1 for r in results if r and r['overall'] == 'PASS') / total * 100:.1f}%",
            ),
            ("Time (seconds)", f"{elapsed:.1f}"),
            ("Speed (URLs/s)", f"{total / elapsed:.1f}"),
            ("User-Agent", ua),
            ("Proxy Used", proxy or "None"),
            ("Input file", input_path),
        ]

        for i, (metric, value) in enumerate(summary_data, start=1):
            ws2.cell(i, 1, metric).font = Font(bold=True, name="Arial")
            ws2.cell(i, 2, value).font = Font(name="Arial")

        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 30

        # Color summary counts
        row_colors = {
            "PASS": fill_pass,
            "WARN": fill_warn,
            "BLOCKED": fill_blocked,
            "ERROR": fill_error,
        }
        for i, (metric, _) in enumerate(summary_data, start=1):
            if metric in row_colors:
                ws2.cell(i, 2).fill = row_colors[metric]

    # ── Final report ──────────────────────────────────────────────────────────
    pass_n = sum(1 for r in results if r and r["overall"] == "PASS")
    warn_n = sum(1 for r in results if r and r["overall"] == "WARN")
    blocked_n = sum(1 for r in results if r and r["overall"] == "BLOCKED")
    error_n = sum(1 for r in results if r and r["overall"] == "ERROR")

    print(f"\n{'='*70}")
    print(f"✓ Done in {elapsed:.1f}s  ({total / elapsed:.1f} URLs/sec)")
    print(f"  PASS:    {pass_n}/{total}")
    print(f"  WARN:    {warn_n}/{total}")
    print(f"  BLOCKED: {blocked_n}/{total}")
    print(f"  ERROR:   {error_n}/{total}")
    print(f"✓ Results saved → {out_path}")
    print(f"{'='*70}\n")


# ══════════════════════════════════════════════════════════════════════════════
# INTERACTIVE MODE
# ══════════════════════════════════════════════════════════════════════════════


def interactive(proxy: str = None):
    print(f"\n{BOLD}Robots Exclusion Checker{RESET}\n")
    url = input("  Enter URL to check: ").strip()
    ua = input("  User-agent [Googlebot]: ").strip() or "Googlebot"
    run(url, ua, proxy)

    again = input("  Check another URL? (y/n): ").strip().lower()
    while again == "y":
        url = input("  URL: ").strip()
        ua = input("  User-agent [Googlebot]: ").strip() or "Googlebot"
        run(url, ua, proxy)
        again = input("  Check another URL? (y/n): ").strip().lower()


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════


if __name__ == "__main__":
    from datetime import datetime

    p = argparse.ArgumentParser(
        description="Robots Exclusion Checker — single URL or bulk Excel mode",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python robots_checker.py                          # interactive\n"
            "  python robots_checker.py https://example.com     # single URL\n"
            "  python robots_checker.py --bulk urls.xlsx        # bulk mode\n"
            "  python robots_checker.py --bulk urls.xlsx --proxy socks5://127.0.0.1:1080\n"
        ),
    )
    p.add_argument("url", nargs="?", help="Single URL to check")
    p.add_argument(
        "user_agent", nargs="?", default="Googlebot", help="User-agent string"
    )
    p.add_argument(
        "--bulk", metavar="FILE.xlsx", help="Bulk mode: Excel file with 'link' column"
    )
    p.add_argument(
        "--ua",
        default="Googlebot",
        help="User-agent for bulk mode (default: Googlebot)",
    )
    p.add_argument(
        "--workers",
        type=int,
        default=20,
        help="Concurrent workers for bulk mode (default: 20)",
    )
    p.add_argument(
        "--out",
        default=None,
        help="Output file for bulk mode (default: results_<timestamp>.xlsx)",
    )
    p.add_argument(
        "--proxy",
        help="Proxy URL (e.g. http://user:pass@host:port or socks5://host:port)",
    )

    args = p.parse_args()

    if args.bulk:
        out = args.out or f"results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        _bulk_mode(args.bulk, args.ua, args.workers, out, args.proxy)
    elif args.url:
        run(args.url, args.user_agent, args.proxy)
    else:
        interactive(args.proxy)
